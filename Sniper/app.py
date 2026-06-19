import streamlit as st
import pandas as pd
import os
import re
import json
import tempfile
from dotenv import load_dotenv
import anthropic
import openpyxl

load_dotenv()

st.set_page_config(
    page_title="Sentiment Sniper",
    page_icon="🎯",
    layout="centered",
    initial_sidebar_state="collapsed",
)

STAR_COLUMN = "rating"
TEXT_COLUMN = "review text"
HEADER_ROW  = 3
SEP_ROW     = 4
DATA_START  = 5
CARD_COLORS = ["#e63946", "#f4a261", "#2a9d8f"]

_S = """<style>
* { font-family: -apple-system, 'Segoe UI', Helvetica, sans-serif; box-sizing: border-box; }
body { margin: 0; padding: 0; background: transparent; }
</style>"""

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.stApp { background: #080808; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 0.5rem !important; padding-bottom: 3rem !important; max-width: 820px !important; }
[data-testid="stFileUploadDropzone"] {
    background: #111 !important; border: 2px dashed #242424 !important; border-radius: 14px !important;
}
[data-testid="stFileUploadDropzone"]:hover { border-color: #e63946 !important; }
[data-testid="stFileUploadDropzone"] p { color: #444 !important; }
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #e63946 0%, #c1121f 100%) !important;
    border: none !important; border-radius: 12px !important;
    font-weight: 600 !important; color: #fff !important;
}
.stButton > button[kind="primary"]:hover { box-shadow: 0 10px 28px rgba(230,57,70,0.38) !important; }
.stAlert { border-radius: 12px !important; }
.stProgress > div > div > div { background: #e63946 !important; }
.stTabs [data-baseweb="tab-list"] { background: transparent; border-bottom: 1px solid #1e1e1e; }
.stTabs [data-baseweb="tab"] { color: #555 !important; }
.stTabs [aria-selected="true"] { color: #fff !important; border-bottom: 2px solid #e63946 !important; }
</style>
""", unsafe_allow_html=True)

# ── SVG ───────────────────────────────────────────────────────────────────────
CROSSHAIR = """<svg xmlns="http://www.w3.org/2000/svg" width="68" height="68"
  viewBox="0 0 100 100" fill="none" stroke="#e63946" stroke-width="5" stroke-linecap="round">
  <circle cx="50" cy="50" r="30"/>
  <circle cx="50" cy="50" r="7" fill="#e63946" stroke="none"/>
  <line x1="50" y1="4"  x2="50" y2="27"/>
  <line x1="50" y1="73" x2="50" y2="96"/>
  <line x1="4"  y1="50" x2="27" y2="50"/>
  <line x1="73" y1="50" x2="96" y2="50"/>
  <line x1="20" y1="20" x2="28" y2="28" stroke-width="3" opacity="0.3"/>
  <line x1="80" y1="20" x2="72" y2="28" stroke-width="3" opacity="0.3"/>
  <line x1="20" y1="80" x2="28" y2="72" stroke-width="3" opacity="0.3"/>
  <line x1="80" y1="80" x2="72" y2="72" stroke-width="3" opacity="0.3"/>
</svg>"""

st.html(f"""{_S}
<div style="text-align:center; padding:2.5rem 0 2rem;">
  {CROSSHAIR}
  <h1 style="color:#fff; font-size:2.3rem; font-weight:800;
             margin:0.65rem 0 0.3rem; letter-spacing:-0.03em;">
    Sentiment Sniper
  </h1>
  <p style="color:#666; font-size:0.93rem; margin:0;">
    Classifica e analizza le recensioni prodotto con l'AI
  </p>
</div>
""")


# ── Data helpers ──────────────────────────────────────────────────────────────
def _parse_stars(value):
    if pd.isna(value): return None
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else None


def load_and_filter(file):
    df = pd.read_excel(file, engine="openpyxl", header=2, skiprows=[3])
    df.columns = df.columns.str.strip().str.lower()
    missing = [c for c in (STAR_COLUMN, TEXT_COLUMN) if c not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti: {missing}. Trovate: {list(df.columns)}")
    df["_stars"] = df[STAR_COLUMN].apply(_parse_stars)
    filtered = df[df["_stars"].isin([1, 2, 3, 4])]
    texts = filtered[TEXT_COLUMN].dropna().astype(str).str.strip().tolist()
    return df, [t for t in texts if t]


# ── API helpers ───────────────────────────────────────────────────────────────
def _client() -> anthropic.Anthropic:
    key = os.getenv("ANTHROPIC_API_KEY")
    if not key:
        raise ValueError("ANTHROPIC_API_KEY non trovata nel file .env")
    return anthropic.Anthropic(api_key=key)


def classify_review(text: str, client: anthropic.Anthropic) -> dict:
    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            messages=[{
                "role": "user",
                "content": (
                    "You are a product review classifier. "
                    "Respond ONLY with a valid JSON object — no explanation, no markdown.\n\n"
                    f'Review: "{text}"\n\n'
                    "Rules:\n"
                    '• "sentiment": exactly one of "Positive", "Neutral", "Negative"\n'
                    '  - Positive = clear satisfaction or praise\n'
                    '  - Neutral  = mixed: some praise AND some complaint\n'
                    '  - Negative = clear dissatisfaction or product failure\n'
                    '• "category": exactly one of "Quality", "Shipping", "Packaging", '
                    '"Usability", "Value", "Durability", "Missing Accessories", "Customer Service"\n'
                    '• "recommended_action": ONE sentence in professional Italian, '
                    'imperative mood, concrete and specific — no filler like '
                    '"si consiglia di", "sarebbe opportuno", "considerare di"\n\n'
                    '{"sentiment":"...","category":"...","recommended_action":"..."}'
                ),
            }],
        )
        raw = msg.content[0].text.strip()
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        return json.loads(m.group() if m else raw)
    except Exception:
        return {"sentiment": "N/A", "category": "N/A", "recommended_action": "N/A"}


def classify_all(filepath: str, on_progress) -> tuple[list[dict], bytes]:
    """
    Classify every row, write Sentiment/Category/Recommended Action into
    columns D-F of the Excel file, and return (results_list, modified_bytes).
    """
    client = _client()
    df     = pd.read_excel(filepath, engine="openpyxl", header=2, skiprows=[3])
    df.columns = df.columns.str.strip().str.lower()

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    for offset, name in enumerate(["Sentiment", "Category", "Recommended Action"]):
        col = 4 + offset
        ws.cell(row=HEADER_ROW, column=col, value=name)
        ws.cell(row=SEP_ROW,    column=col, value=":---")

    results = []
    total   = len(df)

    for i, (_, row) in enumerate(df.iterrows()):
        text = str(row.get(TEXT_COLUMN, "")).strip()
        if text and text.lower() != "nan":
            result = classify_review(text, client)
        else:
            result = {"sentiment": "", "category": "", "recommended_action": ""}

        excel_row = DATA_START + i          # pandas iloc[i] → openpyxl row DATA_START+i
        ws.cell(row=excel_row, column=4, value=result.get("sentiment", ""))
        ws.cell(row=excel_row, column=5, value=result.get("category", ""))
        ws.cell(row=excel_row, column=6, value=result.get("recommended_action", ""))

        results.append({
            "Reviewer":           str(row.get("reviewer", "")),
            "Rating":             str(row.get("rating", "")),
            "Sentiment":          result.get("sentiment", ""),
            "Category":           result.get("category", ""),
            "Recommended Action": result.get("recommended_action", ""),
        })

        on_progress(i + 1, total)

    wb.save(filepath)
    with open(filepath, "rb") as f:
        modified_bytes = f.read()

    return results, modified_bytes


def analyze_aggregate(reviews: list[str]) -> dict:
    client = _client()
    block  = "\n---\n".join(f"Recensione {i+1}: {r}" for i, r in enumerate(reviews))
    prompt = f"""Sei un esperto di analisi delle recensioni prodotto.
Analizza {len(reviews)} recensioni con valutazione 1-4 stelle e identifica i 3 problemi principali ricorrenti.

RECENSIONI:
{block}

Rispondi ESCLUSIVAMENTE con JSON valido, senza markdown né testo extra:
{{"problems":[{{"title":"Titolo breve","relevance":53,"advice":"Consiglio in 1-2 frasi."}},{{"title":"Titolo breve","relevance":40,"advice":"Consiglio in 1-2 frasi."}},{{"title":"Titolo breve","relevance":33,"advice":"Consiglio in 1-2 frasi."}}]}}"""

    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}],
    )
    raw   = msg.content[0].text.strip()
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    return json.loads(match.group() if match else raw)


# ── UI renderers ──────────────────────────────────────────────────────────────
def render_stats(total: int, analyzed: int, coverage: str):
    def card(v, l):
        return (
            f'<div style="flex:1;background:#111;border:1px solid #1e1e1e;'
            f'border-radius:14px;padding:1.3rem 1rem;text-align:center;">'
            f'<div style="font-size:2rem;font-weight:800;color:#e63946;line-height:1;">{v}</div>'
            f'<div style="font-size:0.68rem;color:#3a3a3a;text-transform:uppercase;'
            f'letter-spacing:0.1em;margin-top:0.4rem;">{l}</div></div>'
        )
    st.html(f"""{_S}
<div style="display:flex;gap:1rem;margin:0.5rem 0 1.2rem;">
  {card(total,    "Recensioni totali")}
  {card(analyzed, "Analizzate 1–4 ★")}
  {card(coverage, "Copertura")}
</div>""")


SENTIMENT_ICON = {"Positive": "✅", "Neutral": "🔶", "Negative": "❌", "N/A": "➖"}


def render_results_table(results: list[dict]):
    df = pd.DataFrame(results)
    df["Sentiment"] = df["Sentiment"].apply(
        lambda s: f"{SENTIMENT_ICON.get(s, '')} {s}" if s else s
    )
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Reviewer":           st.column_config.TextColumn("Recensore",   width="small"),
            "Rating":             st.column_config.TextColumn("Voto",        width="small"),
            "Sentiment":          st.column_config.TextColumn("Sentiment",   width="small"),
            "Category":           st.column_config.TextColumn("Categoria",   width="medium"),
            "Recommended Action": st.column_config.TextColumn("Azione consigliata", width="large"),
        },
    )


def render_problem_cards(problems: list[dict]):
    cards = ""
    for i, p in enumerate(problems):
        color = CARD_COLORS[i % len(CARD_COLORS)]
        rel   = p["relevance"]
        cards += (
            f'<div style="background:#111;border:1px solid #1a1a1a;'
            f'border-left:5px solid {color};border-radius:16px;'
            f'padding:1.6rem 1.8rem;margin-bottom:0.9rem;">'
            f'<div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;'
            f'letter-spacing:0.12em;color:{color};margin-bottom:0.45rem;">Problema {i+1}</div>'
            f'<div style="font-size:1.08rem;font-weight:700;color:#fff;'
            f'margin-bottom:1.1rem;line-height:1.4;">{p["title"]}</div>'
            f'<div style="background:#1a1a1a;border-radius:100px;height:6px;'
            f'margin-bottom:0.38rem;overflow:hidden;">'
            f'<div style="width:{rel}%;height:100%;background:{color};border-radius:100px;"></div></div>'
            f'<div style="font-size:0.8rem;color:#3a3a3a;margin-bottom:1.3rem;">'
            f'{rel}% delle recensioni cita questo problema</div>'
            f'<div style="font-size:0.65rem;font-weight:700;text-transform:uppercase;'
            f'letter-spacing:0.12em;color:#2a2a2a;margin-bottom:0.38rem;">Consiglio pratico</div>'
            f'<div style="font-size:0.88rem;color:#999;line-height:1.7;">{p["advice"]}</div>'
            f'</div>'
        )
    st.html(f"""{_S}
<h3 style="color:#fff;font-size:1.05rem;font-weight:700;margin:0.5rem 0 1rem;">
  Problemi principali rilevati
</h3>
{cards}""")


# ── Session state ─────────────────────────────────────────────────────────────
for key, default in [
    ("aggregate_result",   None),
    ("classify_results",   None),
    ("classified_bytes",   None),
    ("last_file",          None),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── File upload ───────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload", type=["xlsx"], label_visibility="collapsed",
    help="Carica il file .xlsx con le recensioni",
)

if not uploaded:
    st.html(f"""{_S}
<div style="text-align:center;color:#2a2a2a;padding:3rem 0;font-size:0.88rem;">
  Trascina qui il file .xlsx oppure clicca per selezionarlo
</div>""")
    st.stop()

# Reset state when a new file is loaded
if uploaded.name != st.session_state.last_file:
    st.session_state.aggregate_result = None
    st.session_state.classify_results = None
    st.session_state.classified_bytes = None
    st.session_state.last_file        = uploaded.name

try:
    df, reviews = load_and_filter(uploaded)
except Exception as e:
    st.error(str(e))
    st.stop()

coverage = f"{round(len(reviews) / len(df) * 100)}%" if len(df) else "0%"
render_stats(len(df), len(reviews), coverage)

if not reviews:
    st.warning("Nessuna recensione 1–4 stelle trovata nel file.")
    st.stop()

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["🔍  Classifica riga per riga", "📊  Report aggregato"])

# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    if st.button("Classifica con AI", key="btn_classify",
                 use_container_width=True, type="primary"):
        # Write uploaded bytes to a temp file so openpyxl can read/write it
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(uploaded.getvalue())
            tmp_path = tmp.name

        prog   = st.progress(0, text="Inizializzazione…")
        status = st.empty()

        def on_progress(done: int, total: int):
            prog.progress(done / total,
                          text=f"Classificando recensione {done} di {total}…")
            status.caption(f"Riga {done}/{total} completata")

        try:
            results, classified_bytes = classify_all(tmp_path, on_progress)
            st.session_state.classify_results  = results
            st.session_state.classified_bytes  = classified_bytes
        except Exception as e:
            st.error(str(e))
        finally:
            os.unlink(tmp_path)

        prog.progress(1.0, text=f"✓ {len(results)} recensioni classificate")
        status.empty()
        st.rerun()

    if st.session_state.classify_results:
        st.success(
            f"✓ {len(st.session_state.classify_results)} recensioni classificate. "
            "Scarica il file aggiornato qui sotto."
        )
        render_results_table(st.session_state.classify_results)

        if st.session_state.classified_bytes:
            st.download_button(
                label="⬇  Scarica Excel con classificazioni",
                data=st.session_state.classified_bytes,
                file_name="reviews_classified.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    if st.button("Genera report aggregato", key="btn_report",
                 use_container_width=True, type="primary"):
        with st.spinner("Analisi in corso…"):
            try:
                st.session_state.aggregate_result = analyze_aggregate(reviews)
            except Exception as e:
                st.error(str(e))

    if st.session_state.aggregate_result:
        render_problem_cards(st.session_state.aggregate_result["problems"])
