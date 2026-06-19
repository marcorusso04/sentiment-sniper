import os
import sys
import re
import json
import pandas as pd
import openpyxl
from dotenv import load_dotenv
import anthropic

load_dotenv()

EXCEL_FILE  = "reviews.xlsx"
STAR_COLUMN = "rating"
TEXT_COLUMN = "review text"

# Excel layout (1-indexed openpyxl rows)
HEADER_ROW = 3   # "Reviewer | Rating | Review Text"
SEP_ROW    = 4   # ":--- | :--- | :---"
DATA_START = 5   # first data row  →  pandas iloc[i] = openpyxl row DATA_START + i


# ── Shared helpers ─────────────────────────────────────────────────────────────

def load_reviews(filepath: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(filepath, engine="openpyxl", header=2, skiprows=[3])
    except FileNotFoundError:
        sys.exit(f"[ERRORE] File '{filepath}' non trovato.")
    except Exception as e:
        sys.exit(f"[ERRORE] Impossibile leggere il file Excel: {e}")
    df.columns = df.columns.str.strip().str.lower()
    missing = [c for c in (STAR_COLUMN, TEXT_COLUMN) if c not in df.columns]
    if missing:
        sys.exit(f"[ERRORE] Colonne mancanti: {missing}. Trovate: {list(df.columns)}")
    return df


def _parse_stars(value):
    if pd.isna(value):
        return None
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else None


def filter_reviews(df: pd.DataFrame) -> list[str]:
    df = df.copy()
    df["_stars"] = df[STAR_COLUMN].apply(_parse_stars)
    filtered = df[df["_stars"].isin([1, 2, 3, 4])]
    texts = filtered[TEXT_COLUMN].dropna().astype(str).str.strip().tolist()
    texts = [t for t in texts if t]
    if not texts:
        sys.exit("[ERRORE] Nessuna recensione 1-4 stelle trovata.")
    return texts


# ── Row-by-row classification ──────────────────────────────────────────────────

def classify_review(text: str, client: anthropic.Anthropic) -> dict:
    """Classify a single review with Haiku (fast, cheap per-row calls)."""
    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            messages=[{
                "role": "user",
                "content": (
                    "You are a product review classifier. "
                    "Respond ONLY with a valid JSON object — no explanation, no markdown.\n\n"
                    f'Review: "{text}"\n\n'
                    "Rules:\n"
                    '• "sentiment": exactly one of "Positive", "Neutral", "Negative"\n'
                    '  - Positive = clear satisfaction or praise\n'
                    '  - Neutral  = mixed: some praise AND some complaint\n'
                    '  - Negative = clear dissatisfaction or product failure\n'
                    '• "category": exactly one of "Quality", "Shipping", "Packaging", '
                    '"Usability", "Value", "Durability", "Missing Accessories", "Customer Service"\n'
                    '• "recommended_action": ONE sentence in professional Italian, '
                    'imperative mood, concrete and specific — no filler like '
                    '"si consiglia di", "sarebbe opportuno", "considerare di"\n\n'
                    '{"sentiment":"...","category":"...","recommended_action":"..."}'
                ),
            }],
        )
        raw = msg.content[0].text.strip()
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        return json.loads(m.group() if m else raw)
    except Exception:
        return {"sentiment": "N/A", "category": "N/A", "recommended_action": "N/A"}


def classify_all(filepath: str, on_progress=None) -> list[dict]:
    """
    Classify every row and write Sentiment / Category / Recommended Action
    into columns D-F of the Excel file without touching existing data.

    on_progress(done: int, total: int) is called after each row (optional).
    Returns a list of result dicts for display.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("[ERRORE] ANTHROPIC_API_KEY non trovata nel file .env")

    df     = load_reviews(filepath)
    client = anthropic.Anthropic(api_key=api_key)
    wb     = openpyxl.load_workbook(filepath)
    ws     = wb.active

    # Write headers + separators for new columns D, E, F
    new_cols = ["Sentiment", "Category", "Recommended Action"]
    for offset, name in enumerate(new_cols):
        col = 4 + offset
        ws.cell(row=HEADER_ROW, column=col, value=name)
        ws.cell(row=SEP_ROW,    column=col, value=":---")

    results = []
    total   = len(df)

    for i, (_, row) in enumerate(df.iterrows()):
        text = str(row.get(TEXT_COLUMN, "")).strip()
        if text and text.lower() != "nan":
            result = classify_review(text, client)
        else:
            result = {"sentiment": "", "category": "", "recommended_action": ""}

        # pandas iloc[i]  →  openpyxl row DATA_START + i
        excel_row = DATA_START + i
        ws.cell(row=excel_row, column=4, value=result.get("sentiment", ""))
        ws.cell(row=excel_row, column=5, value=result.get("category", ""))
        ws.cell(row=excel_row, column=6, value=result.get("recommended_action", ""))

        results.append({
            "Reviewer":           str(row.get("reviewer", "")),
            "Rating":             str(row.get("rating", "")),
            "Sentiment":          result.get("sentiment", ""),
            "Category":           result.get("category", ""),
            "Recommended Action": result.get("recommended_action", ""),
        })

        if on_progress:
            on_progress(i + 1, total)
        else:
            sent = result.get("sentiment", "")
            cat  = result.get("category", "")
            print(f"  [{i+1:2d}/{total}] {sent:<10} | {cat}")

    wb.save(filepath)
    return results


# ── Aggregate report ───────────────────────────────────────────────────────────

def analyze_with_claude(reviews: list[str]) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("[ERRORE] ANTHROPIC_API_KEY non trovata nel file .env")

    client = anthropic.Anthropic(api_key=api_key)
    reviews_block = "\n---\n".join(
        f"Recensione {i+1}: {r}" for i, r in enumerate(reviews)
    )

    prompt = f"""Sei un esperto di analisi delle recensioni di prodotti.
Analizza {len(reviews)} recensioni con valutazione 1-4 stelle e identifica i 3 problemi principali.

RECENSIONI:
{reviews_block}

Rispondi ESCLUSIVAMENTE in questo formato:

PROBLEMA 1: [titolo breve]
RILEVANZA: [X%]
CONSIGLIO: [consiglio pratico in 1-2 frasi]

PROBLEMA 2: [titolo breve]
RILEVANZA: [X%]
CONSIGLIO: [consiglio pratico in 1-2 frasi]

PROBLEMA 3: [titolo breve]
RILEVANZA: [X%]
CONSIGLIO: [consiglio pratico in 1-2 frasi]"""

    print(f"\n[INFO] Invio {len(reviews)} recensioni all'API Anthropic...")
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}],
    )
    return msg.content[0].text


def print_report(analysis: str, total_reviews: int, filtered_count: int) -> None:
    sep = "=" * 62
    print(f"\n{sep}")
    print("          SNIPER — REPORT PROBLEMI RICORRENTI")
    print(sep)
    print(f"  Recensioni totali     : {total_reviews}")
    print(f"  Recensioni 1-4 stelle : {filtered_count}")
    print(sep)
    print()
    print(analysis.strip())
    print()
    print(sep)
    print("  Generato con Claude claude-sonnet-4-6  |  Sniper v2.0")
    print(sep)


# ── CLI entry point ────────────────────────────────────────────────────────────

def main() -> None:
    sep = "=" * 62

    print(f"\n{sep}")
    print("  STEP 1/2 — Classificazione riga per riga  (Haiku)")
    print(sep)
    classify_all(EXCEL_FILE)
    print(f"\n  [OK] Nuove colonne scritte in '{EXCEL_FILE}'\n")

    print(f"{sep}")
    print("  STEP 2/2 — Report aggregato  (Sonnet)")
    print(sep)
    df       = load_reviews(EXCEL_FILE)
    reviews  = filter_reviews(df)
    analysis = analyze_with_claude(reviews)
    print_report(analysis, total_reviews=len(df), filtered_count=len(reviews))


if __name__ == "__main__":
    main()
