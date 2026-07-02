import json
import asyncio
import io
import tempfile
import os

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from file_parser import parse_file, filter_by_rating, COLUMN_ALIASES
from analyzer import classify_reviews, aggregate_by_category, aggregate_by_product, CATEGORIES
from judgeme_client import get_all_reviews as jm_get_reviews, normalize_reviews as jm_normalize
from shopify_reviews_client import get_all_reviews as sp_get_reviews, normalize_reviews as sp_normalize

# ── Brand ─────────────────────────────────────────────────────────────────────
# Colori, font e badge condivisi con landing.html — vedi BRAND_KIT.md

VIOLET = "#7C3AED"
BLUE   = "#3B82F6"
GRAD   = f"linear-gradient(135deg, {VIOLET}, {BLUE})"

CATEGORY_COLORS = {
    "Quality":             "#7C3AED",
    "Shipping":            "#3B82F6",
    "Packaging":           "#F59E0B",
    "Usability":           "#06B6D4",
    "Value":               "#10B981",
    "Durability":          "#EF4444",
    "Missing Accessories": "#EC4899",
    "Customer Service":    "#A78BFA",
}

try:
    with open("assets/icon.svg") as f:
        _ICON_SVG_RAW = f.read()
except FileNotFoundError:
    _ICON_SVG_RAW = ""


def icon_svg(unique_id: str) -> str:
    """Return the icon markup with a unique gradient id, so it can be
    embedded more than once on the same page without duplicate SVG ids."""
    return _ICON_SVG_RAW.replace("sniperGrad", f"sniperGrad-{unique_id}")


ICON_SVG = icon_svg("header")

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Sentiment Sniper",
    page_icon="🎯",
    layout="wide",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Syne:wght@700;800&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    .brand-header { display: flex; align-items: center; gap: 12px; }
    .brand-icon, .sidebar-icon { width: 40px; height: 40px; }
    .sidebar-icon { width: 52px; height: 52px; margin: 0 auto 10px; }
    .brand-icon svg, .sidebar-icon svg { width: 100%; height: 100%; display: block; }
    .brand-title { font-family: 'Syne', sans-serif; font-size: 1.9rem; letter-spacing: -0.02em; }
    .brand-title span { font-weight: 400; color: #F0F0F5; }
    .brand-title strong {
        font-weight: 800;
        background: linear-gradient(135deg, #7C3AED, #3B82F6);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }
    .subtitle { font-size: 0.95rem; color: #8888AA; margin: 4px 0 0 52px; }

    .metric-card {
        background: #13131A; border: 1px solid rgba(255,255,255,0.08);
        border-left: 3px solid #7C3AED; border-radius: 12px;
        padding: 20px; text-align: center;
    }
    .metric-value {
        font-family: 'Syne', sans-serif; font-size: 1.9rem; font-weight: 800;
        background: linear-gradient(135deg, #7C3AED, #3B82F6);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }
    .metric-label { font-size: 0.82rem; color: #8888AA; margin-top: 6px; }

    .stButton>button {
        background: linear-gradient(135deg, #7C3AED, #3B82F6); color: #fff;
        border-radius: 8px; border: none;
        padding: 10px 24px; font-weight: 600;
        box-shadow: 0 4px 20px rgba(124,58,237,0.3);
        transition: opacity 0.2s;
    }
    .stButton>button:hover { opacity: 0.88; color: #fff; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────

st.markdown(f"""
<div class="brand-header">
    <div class="brand-icon">{ICON_SVG}</div>
    <div class="brand-title"><span>Sentiment</span><strong>Sniper</strong></div>
</div>
<div class="subtitle">Analizza le recensioni negative e trasformale in insight azionabili</div>
""", unsafe_allow_html=True)
st.divider()

# ── Session state ─────────────────────────────────────────────────────────────

for key in ["analyzed_reviews", "raw_df", "column_mapping", "mapping_confirmed"]:
    if key not in st.session_state:
        st.session_state[key] = None if key != "mapping_confirmed" else False

# ── Asyncio fix for Streamlit ─────────────────────────────────────────────────

def run_async(coro):
    """Run async coroutine safely inside Streamlit."""
    try:
        loop = asyncio.get_event_loop()
        if loop.is_closed():
            raise RuntimeError
        return loop.run_until_complete(coro)
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()
            asyncio.set_event_loop(None)

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(f'<div class="sidebar-icon">{icon_svg("sidebar")}</div>', unsafe_allow_html=True)
    st.markdown("### Fonte Recensioni")

    source = st.radio(
        "Seleziona fonte",
        options=["📂 Upload CSV/Excel", "⚖️ Judge.me", "🛍️ Shopify Reviews", "🧪 Dati di esempio"],
        label_visibility="collapsed",
    )

    uploaded_file = None
    use_sample    = False
    use_judgeme   = False
    use_shopify   = False

    if source == "📂 Upload CSV/Excel":
        uploaded_file = st.file_uploader(
            "Seleziona file",
            type=["csv", "xlsx", "xls"],
            label_visibility="collapsed",
        )
    elif source == "⚖️ Judge.me":
        st.info("Assicurati di avere **JUDGEME_API_TOKEN** nel tuo `.env`")
        use_judgeme = st.button("🔍 Importa da Judge.me", use_container_width=True)
    elif source == "🛍️ Shopify Reviews":
        st.info("Recupera le recensioni salvate come metafield su Shopify")
        use_shopify = st.button("🔍 Importa da Shopify", use_container_width=True)
    else:
        use_sample = st.button("📂 Carica sample_reviews.csv", use_container_width=True)

    max_stars = st.select_slider(
        "Filtra per stelle (massimo)",
        options=[1, 2, 3, 4],
        value=4,
    )

    st.divider()

    selected_products   = []
    selected_categories = list(CATEGORIES)

    if st.session_state.analyzed_reviews:
        st.markdown("### Filtri Dashboard")
        all_products = sorted(set(r["product_title"] for r in st.session_state.analyzed_reviews))
        selected_products   = st.multiselect("Prodotto", options=all_products, default=all_products)
        selected_categories = st.multiselect("Categoria", options=CATEGORIES, default=CATEGORIES)

        st.divider()
        if st.button("🔄 Nuova Analisi", use_container_width=True):
            for key in ["analyzed_reviews", "raw_df", "column_mapping", "mapping_confirmed"]:
                st.session_state[key] = None if key != "mapping_confirmed" else False
            st.rerun()

# ── Load file & column mapper ─────────────────────────────────────────────────

def load_raw_df(filepath: str) -> pd.DataFrame:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return pd.read_csv(filepath, dtype=str)
    return pd.read_excel(filepath, dtype=str)


def auto_detect_mapping(df: pd.DataFrame) -> dict:
    mapping = {}
    df_cols_lower = {c.lower().strip(): c for c in df.columns}
    for internal, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias.lower() in df_cols_lower:
                mapping[internal] = df_cols_lower[alias.lower()]
                break
    return mapping


if uploaded_file and st.session_state.analyzed_reviews is None and not st.session_state.mapping_confirmed:
    suffix = "." + uploaded_file.name.split(".")[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name
    df = load_raw_df(tmp_path)
    os.unlink(tmp_path)
    st.session_state.raw_df = df
    st.session_state.column_mapping = auto_detect_mapping(df)

# ── Column Mapping UI ─────────────────────────────────────────────────────────

if st.session_state.raw_df is not None and not st.session_state.mapping_confirmed and st.session_state.analyzed_reviews is None:
    df      = st.session_state.raw_df
    mapping = st.session_state.column_mapping

    st.markdown("### 🗂️ Mappa le colonne del tuo file")
    st.caption(f"File caricato: **{len(df)} righe**, **{len(df.columns)} colonne** — `{', '.join(df.columns.tolist())}`")
    st.info("Indica quale colonna del tuo file corrisponde a ciascun campo. I campi con ⚠️ sono obbligatori.")

    col_options = ["— non presente —"] + df.columns.tolist()

    fields = [
        ("review_text",   "📝 Testo Recensione ⚠️", True),
        ("rating",        "⭐ Rating ⚠️",            True),
        ("product_title", "📦 Nome Prodotto",         False),
        ("product_id",    "🔑 ID Prodotto",           False),
        ("reviewer_name", "👤 Nome Reviewer",         False),
        ("date",          "📅 Data",                  False),
    ]

    new_mapping = {}
    cols = st.columns(2)
    for i, (field, label, _) in enumerate(fields):
        current = mapping.get(field, "— non presente —")
        if current not in col_options:
            current = "— non presente —"
        with cols[i % 2]:
            selected = st.selectbox(label, options=col_options, index=col_options.index(current), key=f"map_{field}")
            if selected != "— non presente —":
                new_mapping[field] = selected

    st.markdown("")
    if st.button("✅ Conferma e Avvia Analisi", use_container_width=True):
        if "review_text" not in new_mapping:
            st.error("⚠️ Devi mappare almeno il campo **Testo Recensione**.")
        elif "rating" not in new_mapping:
            st.error("⚠️ Devi mappare almeno il campo **Rating**.")
        else:
            st.session_state.column_mapping = new_mapping
            st.session_state.mapping_confirmed = True
            st.rerun()

    st.stop()

# ── Run analysis ──────────────────────────────────────────────────────────────

def run_analysis_from_df(df: pd.DataFrame, mapping: dict):
    reverse_map = {v: k for k, v in mapping.items()}
    df = df.rename(columns=reverse_map)

    reviews = []
    skipped = 0
    for _, row in df.iterrows():
        row_dict     = row.to_dict()
        review_text  = str(row_dict.get("review_text", "")).strip()
        try:
            rating = int(float(str(row_dict.get("rating", "")).strip()))
        except (ValueError, TypeError):
            skipped += 1
            continue
        if not review_text or review_text.lower() in ("nan", "none", ""):
            skipped += 1
            continue
        if not (1 <= rating <= 5):
            skipped += 1
            continue
        reviews.append({
            "product_id":    str(row_dict.get("product_id", "unknown")).strip(),
            "product_title": str(row_dict.get("product_title", "Unknown Product")).strip(),
            "rating":        rating,
            "review_text":   review_text,
            "reviewer_name": str(row_dict.get("reviewer_name", "Anonymous")).strip(),
            "date":          str(row_dict.get("date", "")).strip() or "N/A",
            "source":        "upload",
        })

    reviews = [r for r in reviews if r["rating"] <= max_stars]
    st.info(f"📋 {len(reviews)} recensioni valide trovate ({skipped} righe saltate)")

    with st.spinner(f"🤖 Analizzo {len(reviews)} recensioni con Claude AI..."):
        analyzed = run_async(classify_reviews(reviews))

    st.session_state.analyzed_reviews = analyzed
    st.success(f"✅ Analisi completata — {len(analyzed)} recensioni processate")
    st.rerun()


def run_analysis_from_file(filepath: str):
    reviews = parse_file(filepath)
    reviews = filter_by_rating(reviews, max_stars=max_stars)
    with st.spinner(f"🤖 Analizzo {len(reviews)} recensioni con Claude AI..."):
        analyzed = run_async(classify_reviews(reviews))
    st.session_state.analyzed_reviews = analyzed
    st.success(f"✅ Analisi completata — {len(analyzed)} recensioni processate")
    st.rerun()


if st.session_state.mapping_confirmed and st.session_state.analyzed_reviews is None and st.session_state.raw_df is not None:
    run_analysis_from_df(st.session_state.raw_df, st.session_state.column_mapping)

if use_sample and st.session_state.analyzed_reviews is None:
    run_analysis_from_file("sample_reviews.csv")

if use_judgeme and st.session_state.analyzed_reviews is None:
    with st.spinner("🔍 Connessione a Judge.me..."):
        raw     = run_async(jm_get_reviews())
        reviews = jm_normalize(raw)
        reviews = filter_by_rating(reviews, max_stars=max_stars)
    if not reviews:
        st.warning("Nessuna recensione trovata su Judge.me. Verifica il token nel .env.")
    else:
        with st.spinner(f"🤖 Analizzo {len(reviews)} recensioni con Claude AI..."):
            analyzed = run_async(classify_reviews(reviews))
        st.session_state.analyzed_reviews = analyzed
        st.success(f"✅ {len(analyzed)} recensioni Judge.me analizzate")
        st.rerun()

if use_shopify and st.session_state.analyzed_reviews is None:
    with st.spinner("🔍 Connessione a Shopify metafields..."):
        raw     = run_async(sp_get_reviews())
        reviews = sp_normalize(raw)
        reviews = filter_by_rating(reviews, max_stars=max_stars)
    if not reviews:
        st.warning("Nessuna recensione trovata nei metafield di Shopify. Il tuo store usa Judge.me o un'altra app?")
    else:
        with st.spinner(f"🤖 Analizzo {len(reviews)} recensioni con Claude AI..."):
            analyzed = run_async(classify_reviews(reviews))
        st.session_state.analyzed_reviews = analyzed
        st.success(f"✅ {len(analyzed)} recensioni Shopify analizzate")
        st.rerun()

# ── Dashboard ─────────────────────────────────────────────────────────────────

if not st.session_state.analyzed_reviews:
    st.info("👈 Carica un file di recensioni dalla sidebar per iniziare l'analisi.")
    st.stop()

reviews = st.session_state.analyzed_reviews
if selected_products:
    reviews = [r for r in reviews if r["product_title"] in selected_products]
if selected_categories:
    reviews = [r for r in reviews if any(c in selected_categories for c in r.get("categories", []))]

by_cat  = aggregate_by_category(reviews)
by_prod = aggregate_by_product(reviews)

# ── KPI ───────────────────────────────────────────────────────────────────────

col1, col2, col3, col4 = st.columns(4)
total_reviews     = len(reviews)
avg_rating        = round(sum(r["rating"] for r in reviews) / total_reviews, 2) if reviews else 0
top_issue         = max(by_cat.items(), key=lambda x: x[1]["count"])[0] if reviews else "—"
products_affected = len(by_prod)

for col, value, label in zip(
    [col1, col2, col3, col4],
    [total_reviews, f"⭐ {avg_rating}", top_issue, products_affected],
    ["Recensioni Analizzate", "Rating Medio", "Problema #1", "Prodotti Coinvolti"],
):
    with col:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{value}</div><div class="metric-label">{label}</div></div>', unsafe_allow_html=True)

st.divider()

# ── Charts ────────────────────────────────────────────────────────────────────

col_left, col_right = st.columns(2)

CHART_LAYOUT = dict(
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    font_color="#F0F0F5",
    legend_font_color="#F0F0F5",
)

with col_left:
    st.markdown("#### 📊 Problemi per Categoria")
    cat_data = pd.DataFrame([
        {"Categoria": cat, "Menzioni": data["count"]}
        for cat, data in by_cat.items() if data["count"] > 0
    ]).sort_values("Menzioni", ascending=True)
    fig_bar = px.bar(cat_data, x="Menzioni", y="Categoria", orientation="h",
                     color="Categoria", color_discrete_map=CATEGORY_COLORS, text="Menzioni")
    fig_bar.update_traces(textposition="outside", textfont_color="#F0F0F5")
    fig_bar.update_layout(
        **CHART_LAYOUT, showlegend=False, height=380, margin=dict(l=0, r=20, t=10, b=0),
        xaxis=dict(gridcolor="rgba(255,255,255,0.08)"),
        yaxis=dict(gridcolor="rgba(255,255,255,0.08)"),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

with col_right:
    st.markdown("#### 🍩 Distribuzione Categorie")
    cat_pie = pd.DataFrame([
        {"Categoria": cat, "Menzioni": data["count"]}
        for cat, data in by_cat.items() if data["count"] > 0
    ])
    fig_pie = px.pie(cat_pie, names="Categoria", values="Menzioni",
                     hole=0.45, color="Categoria", color_discrete_map=CATEGORY_COLORS)
    fig_pie.update_layout(**CHART_LAYOUT, height=380, margin=dict(l=0, r=0, t=10, b=0))
    st.plotly_chart(fig_pie, use_container_width=True)

# ── Heatmap ───────────────────────────────────────────────────────────────────

st.markdown("#### 🔥 Heatmap Prodotti × Categorie")
heatmap_data = []
for product, data in by_prod.items():
    row = {"Prodotto": product}
    for cat in CATEGORIES:
        row[cat] = data.get(cat, 0)
    heatmap_data.append(row)

heatmap_df = pd.DataFrame(heatmap_data).set_index("Prodotto")
fig_heat = go.Figure(data=go.Heatmap(
    z=heatmap_df.values, x=heatmap_df.columns.tolist(), y=heatmap_df.index.tolist(),
    colorscale=[[0, "#13131A"], [0.5, BLUE], [1, VIOLET]],
    text=heatmap_df.values, texttemplate="%{text}", showscale=True,
    xgap=2, ygap=2,
))
fig_heat.update_layout(
    **CHART_LAYOUT, height=max(300, len(by_prod) * 50),
    margin=dict(l=0, r=0, t=10, b=0), xaxis=dict(side="top"),
)
st.plotly_chart(fig_heat, use_container_width=True)

st.divider()

# ── Table ─────────────────────────────────────────────────────────────────────

st.markdown("#### 📋 Recensioni Analizzate")
table_data = [{
    "Prodotto":   r["product_title"],
    "⭐":          r["rating"],
    "Reviewer":   r["reviewer_name"],
    "Data":       r["date"],
    "Categorie":  ", ".join(r.get("categories", [])),
    "Sintesi AI": r.get("summary", ""),
    "Testo":      r["review_text"],
} for r in reviews]
st.dataframe(pd.DataFrame(table_data), use_container_width=True, height=400)
st.divider()

# ── Excel Export ──────────────────────────────────────────────────────────────

def build_excel(reviews, by_cat, by_prod) -> bytes:
    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="7C3AED")
    center      = Alignment(horizontal="center", vertical="center")
    wrap        = Alignment(wrap_text=True, vertical="top")
    thin        = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),  bottom=Side(style="thin", color="D0D0D0"),
    )

    def style_headers(ws, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = thin

    def style_data(ws, wrap_text=True):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = wrap if wrap_text else center
                cell.border = thin

    ws1 = wb.active; ws1.title = "Recensioni"
    ws1.append(["Prodotto", "Rating", "Reviewer", "Data", "Categorie", "Sintesi AI", "Testo", "Fonte"])
    style_headers(ws1, 8)
    for r in reviews:
        ws1.append([r["product_title"], r["rating"], r["reviewer_name"], r["date"],
                    ", ".join(r.get("categories", [])), r.get("summary", ""),
                    r["review_text"], r.get("source", "upload")])
    for i, w in enumerate([35, 8, 20, 12, 35, 45, 60, 10], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    style_data(ws1)
    ws1.freeze_panes = "A2"; ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb.create_sheet("Analisi Categorie")
    ws2.append(["Categoria", "Menzioni Totali", "Prodotti Coinvolti", "Lista Prodotti"])
    style_headers(ws2, 4)
    for cat in CATEGORIES:
        d = by_cat[cat]
        ws2.append([cat, d["count"], len(d["products"]), ", ".join(d["products"])])
    for i, w in enumerate([25, 18, 20, 70], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_data(ws2)
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Analisi Prodotti")
    headers3 = ["Prodotto", "Recensioni", "Rating Medio"] + CATEGORIES
    ws3.append(headers3)
    style_headers(ws3, len(headers3))
    for product, data in by_prod.items():
        ws3.append([product, data["total_reviews"], data["avg_rating"]] +
                   [data.get(cat, 0) for cat in CATEGORIES])
    for i, w in enumerate([35, 12, 14] + [16] * len(CATEGORIES), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.font = Font(name="Arial", size=10); cell.alignment = center; cell.border = thin
            if isinstance(cell.value, int) and cell.value > 0 and cell.column > 3:
                cell.fill = PatternFill("solid", start_color="EDE4FB")
    ws3.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


st.markdown("#### 💾 Esporta Report")
col_a, col_b = st.columns([3, 1])
with col_a:
    st.caption(f"Export di {len(reviews)} recensioni — 3 fogli: Recensioni, Analisi Categorie, Analisi Prodotti.")
with col_b:
    excel_bytes = build_excel(reviews, by_cat, by_prod)
    st.download_button(
        label="⬇️ Download Excel",
        data=excel_bytes,
        file_name="sentiment_sniper_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
